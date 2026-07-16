#!/usr/bin/env python3
# maintained by kewei li
"""
Build a formatted Excel workbook from summary_by_pooling_all.csv (pooling ablation).

- One sheet per (model_type, dataset)
- Main table: pooling + four metrics as mean±std; best value bold, second-best underlined
  (higher better: spearman, pearson, recall_at_k; lower better: rmse)
- To the right: four rank tables (pooling, rank), sorted by ascending rank
- Sheet ``avg_rank_4combos``: mean (and std across combos) of per-sheet ranks for
  each pooling over the four (model_type, dataset) groups; writes
  ``avg_rank_across_4_combos.csv`` beside the workbook.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]


def _default_csv_path() -> Path:
    return (
        REPO_ROOT
        / "outputs"
        / "ablation-new-data"
        / "_pooling_stats"
        / "summary_by_pooling_all.csv"
    ).resolve()


def _default_out_path() -> Path:
    return (
        REPO_ROOT
        / "outputs"
        / "ablation-new-data"
        / "_pooling_stats"
        / "pooling_ablation_stats.xlsx"
    ).resolve()


def _fmt_pm(mean: float, std: float, mean_digits: int = 4, std_digits: int = 4) -> str:
    return f"{mean:.{mean_digits}f}±{std:.{std_digits}f}"


def _best_second(
    values: List[float], higher_is_better: bool
) -> Tuple[Optional[float], Optional[float]]:
    """Return (best_value, second_distinct_value) for formatting."""
    if not values:
        return None, None
    uniq = sorted(set(values), reverse=higher_is_better)
    best = uniq[0]
    second = uniq[1] if len(uniq) > 1 else None
    return best, second


def _rank_series(
    s: pd.Series, higher_is_better: bool
) -> pd.Series:
    """Rank 1 = best; ties get same rank (min)."""
    return s.rank(ascending=not higher_is_better, method="min").astype(int)


_RANK_METRIC_SPECS: List[Tuple[str, str, bool]] = [
    ("spearman", "spearman_mean", True),
    ("pearson", "pearson_mean", True),
    ("recall_at_k", "recall_at_k_mean", True),
    ("rmse", "rmse_mean", False),
]


def _std_ddof1(s: pd.Series) -> float:
    return float(s.std(ddof=1)) if len(s) > 1 else 0.0


def average_ranks_across_combos(df: pd.DataFrame) -> pd.DataFrame:
    """
    For each pooling, mean and std (ddof=1) of ranks across the four
    (model_type, dataset) groups. Same ranking as per-sheet rank tables.
    Adds overall_avg_rank = mean of the four average ranks (lower is better).
    """
    parts: List[pd.DataFrame] = []
    for (_model, _dataset), grp in df.groupby(["model_type", "dataset"], sort=False):
        sub = grp.set_index("pooling")
        block = pd.DataFrame({"pooling": sub.index.tolist()})
        for key, colname, higher in _RANK_METRIC_SPECS:
            block[f"rank_{key}"] = _rank_series(sub[colname], higher).values
        parts.append(block)
    if not parts:
        return pd.DataFrame()

    stacked = pd.concat(parts, ignore_index=True)
    agg_kw = {}
    for key, _, _ in _RANK_METRIC_SPECS:
        col = f"rank_{key}"
        agg_kw[f"avg_rank_{key}"] = pd.NamedAgg(column=col, aggfunc="mean")
        agg_kw[f"std_rank_{key}"] = pd.NamedAgg(column=col, aggfunc=_std_ddof1)

    out = stacked.groupby("pooling", as_index=False).agg(**agg_kw)
    avg_cols = [f"avg_rank_{k}" for k, _, _ in _RANK_METRIC_SPECS]
    out["overall_avg_rank"] = out[avg_cols].mean(axis=1)
    out = out.sort_values(["overall_avg_rank"] + avg_cols).reset_index(drop=True)
    return out


def _write_avg_rank_sheet(ws, avg_df: pd.DataFrame, recall_topk: int) -> None:
    """pooling + avg/std per metric + overall_avg_rank; bold best overall."""
    font_bold = Font(bold=True)
    if avg_df.empty:
        ws.cell(row=1, column=1, value="(no data)")
        return

    rk = int(recall_topk)
    headers = [
        "pooling",
        "avg_rank_spearman",
        "std_rank_spearman",
        "avg_rank_pearson",
        "std_rank_pearson",
        f"avg_rank_recall@{rk}",
        f"std_rank_recall@{rk}",
        "avg_rank_rmse",
        "std_rank_rmse",
        "overall_avg_rank",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font_bold
        c.alignment = Alignment(horizontal="center")

    best_overall = float(avg_df["overall_avg_rank"].min())

    for ridx, row in enumerate(avg_df.itertuples(index=False), start=2):
        vals = [
            row.pooling,
            round(float(row.avg_rank_spearman), 4),
            round(float(row.std_rank_spearman), 4),
            round(float(row.avg_rank_pearson), 4),
            round(float(row.std_rank_pearson), 4),
            round(float(row.avg_rank_recall_at_k), 4),
            round(float(row.std_rank_recall_at_k), 4),
            round(float(row.avg_rank_rmse), 4),
            round(float(row.std_rank_rmse), 4),
            round(float(row.overall_avg_rank), 4),
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=ridx, column=col, value=v)
            if col == len(vals) and abs(float(v) - best_overall) < 1e-9:
                cell.font = font_bold

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18 if col > 1 else 22


def _write_sheet(
    ws,
    sub: pd.DataFrame,
    recall_label: str = "recall@50",
) -> None:
    font_bold = Font(bold=True)
    font_under = Font(underline="single")

    # Row 1: main headers (A–E) + rank table titles (G+). Row 2: rank subheaders only;
    # main data starts row 3 so it does not overlap rank "pooling|rank" headers.
    data_start_row = 3

    headers = [
        "pooling",
        "spearman (mean±std)",
        "pearson (mean±std)",
        f"{recall_label} (mean±std)",
        "RMSE (mean±std)",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font_bold
        c.alignment = Alignment(horizontal="center")

    spe_vals = sub["spearman_mean"].tolist()
    pea_vals = sub["pearson_mean"].tolist()
    rec_vals = sub["recall_at_k_mean"].tolist()
    rmse_vals = sub["rmse_mean"].tolist()

    b_sp, s_sp = _best_second(spe_vals, True)
    b_pe, s_pe = _best_second(pea_vals, True)
    b_re, s_re = _best_second(rec_vals, True)
    b_rm, s_rm = _best_second(rmse_vals, False)

    for ridx, row in enumerate(sub.itertuples(index=False), start=data_start_row):
        ws.cell(row=ridx, column=1, value=row.pooling)

        mean, std = row.spearman_mean, row.spearman_std_ddof1
        cell = ws.cell(row=ridx, column=2, value=_fmt_pm(mean, std))
        if b_sp is not None and mean == b_sp:
            cell.font = font_bold
        elif s_sp is not None and mean == s_sp:
            cell.font = font_under

        mean, std = row.pearson_mean, row.pearson_std_ddof1
        cell = ws.cell(row=ridx, column=3, value=_fmt_pm(mean, std))
        if b_pe is not None and mean == b_pe:
            cell.font = font_bold
        elif s_pe is not None and mean == s_pe:
            cell.font = font_under

        mean, std = row.recall_at_k_mean, row.recall_at_k_std_ddof1
        cell = ws.cell(row=ridx, column=4, value=_fmt_pm(mean, std, mean_digits=2, std_digits=2))
        if b_re is not None and mean == b_re:
            cell.font = font_bold
        elif s_re is not None and mean == s_re:
            cell.font = font_under

        mean, std = row.rmse_mean, row.rmse_std_ddof1
        cell = ws.cell(row=ridx, column=5, value=_fmt_pm(mean, std))
        if b_rm is not None and mean == b_rm:
            cell.font = font_bold
        elif s_rm is not None and mean == s_rm:
            cell.font = font_under

    # Rank tables start at column G (7); column F left blank as spacer
    rank_start_col = 7
    blocks = [
        ("Spearman rank", "spearman_mean", True),
        ("Pearson rank", "pearson_mean", True),
        (f"{recall_label} rank", "recall_at_k_mean", True),
        ("RMSE rank", "rmse_mean", False),
    ]

    for bi, (title, colname, higher) in enumerate(blocks):
        base = rank_start_col + bi * 3  # two cols + one spacer between blocks
        ws.merge_cells(
            start_row=1,
            start_column=base,
            end_row=1,
            end_column=base + 1,
        )
        tcell = ws.cell(row=1, column=base, value=title)
        tcell.font = Font(bold=True)
        tcell.alignment = Alignment(horizontal="center")

        ws.cell(row=2, column=base, value="pooling").font = font_bold
        ws.cell(row=2, column=base + 1, value="rank").font = font_bold

        ranks = _rank_series(sub.set_index("pooling")[colname], higher)
        rank_df = (
            ranks.reset_index()
            .rename(columns={colname: "rank"})
            .sort_values(["rank", "pooling"])
            .reset_index(drop=True)
        )
        for i, pr in enumerate(rank_df.itertuples(index=False), start=data_start_row):
            ws.cell(row=i, column=base, value=pr.pooling)
            ws.cell(row=i, column=base + 1, value=int(pr.rank))

    # Column widths (rough)
    for col in range(1, rank_start_col + len(blocks) * 3):
        letter = get_column_letter(col)
        if col == 1:
            ws.column_dimensions[letter].width = 22
        elif col <= 5:
            ws.column_dimensions[letter].width = 18
        elif col >= rank_start_col:
            ws.column_dimensions[letter].width = 14


def build_workbook(
    df: pd.DataFrame,
    out_path: Path,
    recall_topk: Optional[int] = None,
) -> Path:
    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    k = int(df["recall_topk"].iloc[0]) if recall_topk is None else int(recall_topk)

    for (model, dataset), grp in df.groupby(["model_type", "dataset"], sort=False):
        name = f"{model}_{dataset}"
        if len(name) > 31:
            name = name[:31]
        ws = wb.create_sheet(title=name)
        recall_label = f"recall@{k}"
        _write_sheet(ws, grp.reset_index(drop=True), recall_label=recall_label)

    avg_df = average_ranks_across_combos(df)
    avg_csv = out_path.parent / "avg_rank_across_4_combos.csv"
    if not avg_df.empty:
        avg_df.to_csv(avg_csv, index=False)
    ws_avg = wb.create_sheet(title="avg_rank_4combos")
    _write_avg_rank_sheet(ws_avg, avg_df, k)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return avg_csv


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--input-csv",
        type=Path,
        default=None,
        help=f"summary_by_pooling_all.csv (default: {_default_csv_path()})",
    )
    ap.add_argument(
        "--output-xlsx",
        type=Path,
        default=None,
        help=f"Output .xlsx (default: {_default_out_path()})",
    )
    args = ap.parse_args()

    csv_path = (
        args.input_csv.expanduser().resolve()
        if args.input_csv is not None
        else _default_csv_path()
    )
    out_path = (
        args.output_xlsx.expanduser().resolve()
        if args.output_xlsx is not None
        else _default_out_path()
    )

    if not csv_path.is_file():
        print(f"Error: input not found: {csv_path}", file=sys.stderr)
        return 1

    df = pd.read_csv(csv_path)
    required = [
        "model_type",
        "dataset",
        "pooling",
        "spearman_mean",
        "spearman_std_ddof1",
        "pearson_mean",
        "pearson_std_ddof1",
        "recall_at_k_mean",
        "recall_at_k_std_ddof1",
        "rmse_mean",
        "rmse_std_ddof1",
        "recall_topk",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"Error: CSV missing columns: {missing}", file=sys.stderr)
        return 1

    avg_csv = build_workbook(df, out_path)
    print(f"Wrote {out_path}")
    if avg_csv.is_file():
        print(f"Wrote {avg_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
